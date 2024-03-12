from io import BytesIO

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import frappe


class ExcelExporter:
    def __init__(self):
        self.wb = openpyxl.Workbook()

    def create_sheet(self, **kwargs):
        """
        create worksheet
        :param sheet_name - name for the worksheet
        :param filters - A data dictionary to added in sheet
        :param merged_headers - A dict of List
            @example: {
                'label': [column1, colum2]
            }
        :param headers: A List of dictionary (cell properties will be optional)
        :param data: A list of dictionary to append data to sheet
        """

        Worksheet().create(workbook=self.wb, **kwargs)

    def save_workbook(self, file_name=None):
        """Save workbook"""
        if file_name:
            self.wb.save(file_name)
            return self.wb

        xlsx_file = BytesIO()
        self.wb.save(xlsx_file)
        return xlsx_file

    def remove_sheet(self, sheet_name):
        """Remove worksheet"""
        if sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[sheet_name])

    def export(self, file_name):
        # write out response as a xlsx type
        if file_name[-4:] != ".xlsx":
            file_name = f"{file_name}.xlsx"

        xlsx_file = self.save_workbook()

        frappe.local.response["filename"] = file_name
        frappe.local.response["filecontent"] = xlsx_file.getvalue()
        frappe.local.response["type"] = "binary"


class Worksheet:
    data_format = frappe._dict(
        {
            "font_family": "Calibri",
            "font_size": 9,
            "bold": False,
            "horizontal": "general",
            "number_format": "General",
            "width": 20,
            "height": 20,
            "vertical": "center",
            "wrap_text": False,
            "bg_color": "f2f2f2",
        }
    )
    filter_format = data_format.copy().update({"bg_color": None, "bold": True})
    header_format = frappe._dict(
        {
            "font_family": "Calibri",
            "font_size": 9,
            "bold": True,
            "horizontal": "center",
            "number_format": "General",
            "width": 20,
            "height": 30,
            "vertical": "center",
            "wrap_text": True,
            "bg_color": "d9d9d9",
        }
    )
    default_styles = frappe._dict(
        {
            "is_header": "header_format",
            "is_total": "header_format",
            "is_data": "data_format",
            "is_filter": "filter_format",
        }
    )

    def __init__(self):
        self.row_dimension = 1
        self.column_dimension = 1

    def create(
        self,
        workbook,
        sheet_name,
        headers,
        data,
        filters=None,
        merged_headers=None,
        add_totals=True,
    ):
        """Create worksheet"""
        self.headers = headers

        self.ws = workbook.create_sheet(sheet_name)
        self.add_data(filters, is_filter=True)
        self.add_merged_header(merged_headers)
        self.add_data(headers, is_header=True)
        self.add_data(data, is_data=True)

        if add_totals:
            self.add_data(self.get_totals(), is_total=True)

        self.apply_conditional_formatting(add_totals)

    def add_data(self, data, **kwargs):
        if not data:
            return

        if kwargs.get("is_data"):
            self.data_row = self.row_dimension

        for row in self.parse_data(data):
            for idx, val in enumerate(row, 1):
                cell = self.ws.cell(row=self.row_dimension, column=idx)
                self.apply_format(row=self.row_dimension, column=idx, **kwargs)
                cell.value = val

            self.row_dimension += 1

    def add_merged_header(self, merged_headers):
        if not merged_headers:
            return

        for key, value in merged_headers.items():
            merge_from_idx = self.get_column_index(value[0])
            merge_to_idx = self.get_column_index(value[1])

            cell_range = self.get_range(
                start_row=self.row_dimension,
                start_column=merge_from_idx,
                end_row=self.row_dimension,
                end_column=merge_to_idx,
            )

            self.ws.cell(row=self.row_dimension, column=merge_from_idx).value = key

            self.ws.merge_cells(cell_range)

            self.apply_format(
                row=self.row_dimension,
                column=merge_from_idx,
                is_header=True,
            )

        self.row_dimension += 1

    def get_totals(self):
        """build total row array of fields to be calculated"""
        total_row = []

        for idx, column in enumerate(self.headers, 1):
            if idx == 1:
                total_row.append("Totals")
            elif column.get("fieldtype") in ("Float", "Int"):
                cell_range = self.get_range(self.data_row, idx, self.ws.max_row, idx)
                total_row.append(f"=SUM({cell_range})")
            else:
                total_row.append("")

        return total_row

    def apply_format(self, row, column, **kwargs):
        """Get style if defined or apply default format to the cell"""

        key, value = kwargs.popitem()
        if not value:
            return

        # get default style
        style_name = self.default_styles.get(key)
        style = getattr(self, style_name).copy()

        # update custom style
        custom_styles = self.headers[column - 1].get(style_name)
        if custom_styles:
            style.update(custom_styles)

        if key == "is_total":
            style.update(
                {
                    "horizontal": self.data_format.horizontal,
                    "height": self.data_format.height,
                }
            )

        self.apply_style(row, column, style)

    def apply_style(self, row, column, style):
        """Apply style to cell"""

        cell = self.ws.cell(row=row, column=column)
        cell.font = Font(name=style.font_family, size=style.font_size, bold=style.bold)
        cell.alignment = Alignment(
            horizontal=style.horizontal,
            vertical=style.vertical,
            wrap_text=style.wrap_text,
        )
        cell.number_format = style.number_format

        if style.bg_color:
            cell.fill = PatternFill(fill_type="solid", fgColor=style.bg_color)

        self.ws.column_dimensions[get_column_letter(column)].width = style.width
        self.ws.row_dimensions[row].height = style.height

    def apply_conditional_formatting(self, has_totals):
        """Apply conditional formatting to data based on comparable fields as defined in headers"""

        for row in self.headers:
            if not (compare_field := row.get("compare_with")):
                continue

            column = get_column_letter(self.get_column_index(row["fieldname"]))
            compare_column = get_column_letter(self.get_column_index(compare_field))

            # eg formula used: IF(ISBLANK(H6), FALSE, H6<>R6)
            formula = f"IF(ISBLANK({column}{self.data_row}), FALSE, {column}{self.data_row}<>{compare_column}{self.data_row})"

            cell_range = self.get_range(
                start_row=self.data_row,
                start_column=column,
                end_row=self.ws.max_row - has_totals,
                end_column=column,
            )

            self.ws.conditional_formatting.add(
                cell_range,
                FormulaRule(
                    formula=[formula],
                    stopIfTrue=True,
                    font=Font(
                        name=self.data_format.get("font_family"),
                        size=self.data_format.get("font_size"),
                        bold=True,
                        color="FF0000",
                    ),
                ),
            )

    def parse_data(self, data):
        """Convert data to List of Lists"""
        out = []

        if isinstance(data, dict):
            for key, value in data.items():
                # eg: {"fieldname": "value"} => ["fieldname", "value"]. for filters.
                out.append([key, value])

        elif isinstance(data, list):
            # eg: ["value1", "value2"] => ["value1", "value2"]. for totals.
            if isinstance(data[0], str):
                return [data]

            for row in data:
                # eg: [{"label": "value1"}] => "value1". for headers.
                if row.get("label"):
                    out.append(row.get("label"))
                else:
                    # eg: [{"fieldname1": "value1", "fieldname2": "value2"}] => ["value1", "value2"]. for data.
                    out.append([row.get(field["fieldname"]) for field in self.headers])

            if row.get("label"):
                return [out]

        return out

    def get_range(self, start_row, start_column, end_row, end_column, freeze=False):
        """
        Get range of cells
        parameters:
            start_row (int): row number of the first cell
            start_column (int | string): column number / letter of the first cell
            end_row (int): row number of the last cell
            end_column (int | string): column number / letter of the last cell
            freeze (bool): freeze the range

        returns:
            string: range of cells eg: A1:B2
        """

        if isinstance(start_column, int):
            start_column = get_column_letter(start_column)

        if isinstance(end_column, int):
            end_column = get_column_letter(end_column)

        if freeze:
            return f"${start_column}${start_row}:${end_column}${end_row}"

        return f"{start_column}{start_row}:{end_column}{end_row}"

    def get_column_index(self, column_name):
        """Get column index / position from column name"""

        for idx, field in enumerate(self.headers, 1):
            if field["fieldname"] == column_name:
                return idx
