from io import BytesIO

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import frappe


class ExcelExporter:
    def __init__(self):
        self.wb = openpyxl.Workbook()

    def create_sheet(self, **kwargs):
        """
        create worksheet
        :param sheet_name - name for the worksheet
        :param filters - A data dictionary to added in sheet
        :param merged_headers - A dict of List
            @example: {
                'label': [column1, colum2]
            }
        :param headers: A List of dictionary (cell properties will be optional)
        :param data: A list of dictionary to append data to sheet
        """

        Worksheet().create(workbook=self.wb, **kwargs)

    def save_workbook(self, file_name=None):
        """Save workbook"""
        if file_name:
            self.wb.save(file_name)
            return self.wb

        xlsx_file = BytesIO()
        self.wb.save(xlsx_file)
        return xlsx_file

    def remove_sheet(self, sheet_name):
        """Remove worksheet"""
        if sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[sheet_name])

    def export(self, file_name):
        # write out response as a xlsx type
        if file_name[-4:] != ".xlsx":
            file_name = f"{file_name}.xlsx"

        xlsx_file = self.save_workbook()

        frappe.local.response["filename"] = file_name
        frappe.local.response["filecontent"] = xlsx_file.getvalue()
        frappe.local.response["type"] = "binary"


class Worksheet:
    def __init__(self):
        self.row_dimension = 1
        self.column_dimension = 1

    def create(
        self,
        workbook,
        sheet_name,
        headers,
        data,
        filters=None,
        merged_headers=None,
        add_totals=True,
    ):
        """Create worksheet"""
        self.headers = headers

        self.ws = workbook.create_sheet(sheet_name)
        self.add_filters(filters)
        self.add_merged_header(merged_headers)
        self.add_headers(headers)
        self.add_data(data)

        if add_totals:
            self.add_total_row()

        self.apply_conditional_formatting()

    def add_filters(self, filter_data):
        if not filter_data:
            return

        parsed_data = self.parse_data(filter_data)

        for i, row in enumerate(parsed_data, 1):
            for j, val in enumerate(row):
                cell = self.ws.cell(row=self.row_dimension, column=j + 1)
                self.apply_style(
                    self.row_dimension,
                    j + 1,
                    column_index=j,
                )
                cell.value = val

            self.row_dimension += 1

    def add_merged_header(self, merged_headers):
        if not merged_headers:
            return

        for key, value in merged_headers.items():
            merge_from_idx = self.get_column_index(value[0])
            merge_to_idx = self.get_column_index(value[1])

            range = self.get_range(
                start_row=self.row_dimension,
                start_column=merge_from_idx,
                end_row=self.row_dimension,
                end_column=merge_to_idx,
            )

            self.ws.cell(row=self.row_dimension, column=merge_from_idx).value = key

            self.ws.merge_cells(range)

            self.apply_style(
                self.row_dimension,
                merge_from_idx,
                column_index=merge_from_idx,
                is_header=True,
            )

        self.row_dimension += 1

    def add_headers(self, header_data):
        if not header_data:
            return

        parsed_data = self.parse_data(header_data, is_header=True)

        for i, row in enumerate(parsed_data, 1):
            for j, val in enumerate(row):
                cell = self.ws.cell(row=self.row_dimension, column=j + 1)

                self.apply_style(
                    self.row_dimension, j + 1, column_index=j, is_header=True
                )
                cell.value = val

            self.row_dimension += 1

    def add_data(self, data):
        if not data:
            return

        self.data_row = self.row_dimension

        parsed_data = self.parse_data(data)

        for i, row in enumerate(parsed_data, 1):
            for j, val in enumerate(row):
                cell = self.ws.cell(row=self.row_dimension, column=j + 1)

                self.apply_style(
                    self.row_dimension, j + 1, column_index=j, is_data=True
                )
                cell.value = val

            self.row_dimension += 1

    def add_total_row(self):
        """Add total row to the sheet"""

        total_row = self.get_totals()

        parsed_data = self.parse_data(total_row)

        for i, row in enumerate(parsed_data, 1):
            for j, val in enumerate(row):
                cell = self.ws.cell(row=self.row_dimension, column=j + 1)

                self.apply_style(
                    self.row_dimension,
                    j + 1,
                    column_index=j,
                    is_header=True,
                    is_total=True,
                )
                cell.value = val

            self.row_dimension += 1

    def get_totals(self):
        """build total row array of fields to be calculated"""
        total_row = []

        for idx, property in enumerate(self.headers, 1):
            if idx == 1:
                total_row.append("Totals")
            elif property.get("fieldtype") in ("Float", "Int"):
                range = self.get_range(self.data_row, idx, self.ws.max_row, idx)
                total_row.append(f"=SUM({range})")
            else:
                total_row.append("")

        return total_row

    def apply_style(
        self, row, column, column_index, is_header=False, is_data=False, is_total=False
    ):
        """Apply style to cell"""
        self.get_properties(column_index, is_header, is_data, is_total)

        cell = self.ws.cell(row=row, column=column)
        cell.font = Font(
            name=self.font_family,
            size=self.font_size,
            bold=self.bold,
        )
        cell.alignment = Alignment(
            horizontal=self.horizontal_align,
            vertical=self.vertical_align,
            wrap_text=self.wrap_text,
        )
        cell.number_format = self.format if self.format else "General"

        if self.bg_color:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.bg_color,
            )
        self.ws.column_dimensions[get_column_letter(column)].width = self.width
        self.ws.row_dimensions[row].height = self.height

    def apply_conditional_formatting(self):
        """Apply conditional formatting to cell"""

        for i, row in enumerate(self.headers, 1):
            if "formula" not in row:
                continue

            end_column = (
                self.get_column_index(row.get("compare_field")) or self.ws.max_column
            )

            range = self.get_range(
                start_row=self.data_row,
                start_column=self.get_column_index(row["fieldname"]),
                end_row=self.ws.max_row,
                end_column=end_column - 1,
            )

            self.ws.conditional_formatting.add(
                range,
                FormulaRule(
                    formula=[self.formula],
                    stopIfTrue=True,
                    font=Font(
                        name=self.font_family,
                        size=self.font_size,
                        bold=self.bold,
                        color="FF0000",
                    ),
                ),
            )

    def get_properties(
        self, column=None, is_header=False, is_data=False, is_total=False
    ):
        """Get all properties defined in a header for cell"""

        properties = self.headers[column]

        self.font_family = "Calibri"
        self.font_size = 9
        self.bold = True
        self.align_header = "center"
        self.align_data = "general"
        self.format = "General"
        self.width = 20
        self.height = 20
        self.vertical_align = "bottom"
        self.wrap_text = False
        self.bg_color = None

        for (k, v) in properties.items():
            setattr(self, k, v)

        if is_header:
            self.horizontal_align = self.align_header
            self.bg_color = self.bg_color
            self.font_size = self.font_size
            self.vertical_align = "center"
            self.wrap_text = True
            self.height = 30
            if is_total:
                self.horizontal_align = self.align_data
                self.height = 20
        elif is_data:
            self.horizontal_align = self.align_data
            self.bg_color = self.bg_color_data
            self.bold = False
        else:
            self.horizontal_align = self.align_data
            self.bg_color = None

    def parse_data(self, data, is_header=False):
        """Convert data to List of Lists"""
        csv_list = []

        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, list):
                    # To get keys from Merged headers having keys and list of columns to merge
                    csv_list.append(list(data.keys()))
                    return csv_list
                else:
                    # To get filters value key as Label and it's value
                    csv_list.append([key, value])

        if isinstance(data, list):
            # For Headers and Data
            csv_list = self.build_csv_array(data, is_header)

        return csv_list

    def build_csv_array(self, data, is_header=False):
        """Builds a csv data array"""
        csv_array = []
        csv_row = []

        for row in data:
            if isinstance(row, dict):
                if is_header:
                    # Fetch label from list of dict from headers
                    csv_row.append(row.get("label"))
                else:
                    # Fetch only values from list of dictionary
                    csv_array.append(list(row.values()))
            else:
                # If it's a single list with element only
                csv_array.append(data)
                break

        # Only append to csv_array if csv_row has multiple values
        if len(csv_row) >= 1:
            csv_array.append(csv_row)

        return csv_array

    def get_range(self, start_row, start_column, end_row, end_column, freeze=False):
        start_column_letter = get_column_letter(start_column)
        end_column_letter = get_column_letter(end_column)

        if freeze:
            return f"${start_column_letter}${start_row}:${end_column_letter}${end_row}"

        return f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    def get_column_index(self, column_name):
        """Get column index from column name"""
        for (idx, field) in enumerate(self.headers, 1):
            if field["fieldname"] == column_name:
                return idx
