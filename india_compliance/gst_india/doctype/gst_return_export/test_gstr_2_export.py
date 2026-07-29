# Copyright (c) 2026, Resilient Tech and contributors
# For license information, please see license.txt

"""
Tests for the GSTR-2A / 2B Excel exporter (gstr_2_export.py): per-cell value assertions
for the tricky logic, plus committed golden workbooks (gst_india/data) as regression nets.
Mocked unit builds for coverage, one integration test through build_export.
"""

from contextlib import ExitStack, contextmanager
from io import BytesIO
from typing import ClassVar
from unittest.mock import Mock, patch
from zipfile import ZipFile

import frappe
import openpyxl
from frappe import parse_json, read_file
from frappe.tests import IntegrationTestCase
from frappe.utils import add_days, now_datetime

from india_compliance.gst_india.doctype.gst_return_export.gstr_2_export import (
    DOCTYPE,
    GovReturnExporter,
    GSTR2AExporter,
    GSTR2BExporter,
    _group_periods,
    build_export,
    delete_stale_export_files,
    download_export_file,
)
from india_compliance.gst_india.doctype.gst_return_log.gst_return_log import (
    store_raw_return_data,
)
from india_compliance.gst_india.utils import get_data_file_path
from india_compliance.gst_india.utils.gstr_utils import ReturnType

EXPORT_MODULE = "india_compliance.gst_india.doctype.gst_return_export.gstr_2_export"

GSTIN_2B = "01AABCE2207R1Z5"
PERIOD_2B = "032020"
GOLDEN_2B = "test_gstr2b_export_golden.xlsx"

GSTIN_2A = "24AACT1234F1Z5"
PERIOD_2A = "052024"
GOLDEN_2A = "test_gstr2a_export_golden.xlsx"

MOCK_LEGAL = "LEGAL NAME"
MOCK_TRADE = "TRADE NAME"
MOCK_BUSINESS = "GSTN"

RAW_2A = {
    "b2b": [
        {
            "ctin": GSTIN_2A,
            "trdnm": "Acme Traders",
            "inv": [
                {
                    "inum": "INV-001",
                    "idt": "10-05-2024",
                    "val": 14200,
                    "pos": "06",
                    "itms": [
                        {"itm_det": {"rt": 18, "txval": 10000, "iamt": 0, "camt": 900, "samt": 900}},
                        {"itm_det": {"rt": 5, "txval": 2000, "iamt": 100}},
                    ],
                }
            ],
        }
    ],
    "isd": [
        {
            "ctin": "16DEFPS8555D1Z7",
            "trdnm": "ISD Co",
            "doclist": [
                {
                    "isd_docty": "ISD",
                    "docnum": "S0080",
                    "docdt": "03-03-2016",
                    "iamt": 20,
                    "camt": 20,
                    "samt": 20,
                },
                {
                    "isd_docty": "ISDCN",
                    "docnum": "CN-1",
                    "docdt": "04-03-2016",
                    "iamt": 5,
                    "camt": 5,
                    "samt": 5,
                },
            ],
        }
    ],
    "tcs": [{"etin": "27AAECS1234F1Z5", "sup_val": 1000, "tx_val": 900, "iamt": 10, "camt": 5, "samt": 5}],
}


def _load(content):
    return openpyxl.load_workbook(BytesIO(content))


def _cells(ws):
    """{(row, col): value} for every non-empty cell — the comparable content of a sheet."""
    return {
        (cell.row, cell.column): cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    }


def _data_row_count(ws, start=7, col=1):
    """Number of data rows (non-empty first column) below the header block."""
    return sum(1 for r in range(start, ws.max_row + 1) if ws.cell(r, col).value not in (None, ""))


@contextmanager
def _mock_names(raw):
    """Feed the raw payload and mock registry name lookups deterministically."""
    with ExitStack() as stack:
        stack.enter_context(patch(f"{EXPORT_MODULE}.get_raw_return_data", return_value=raw))
        stack.enter_context(
            patch.object(
                GovReturnExporter,
                "_gstin_info",
                staticmethod(
                    lambda gstin: {
                        "business_name": MOCK_BUSINESS,
                        "legal_name": MOCK_LEGAL,
                        "trade_name": MOCK_TRADE,
                    }
                ),
            )
        )
        yield stack


def _build_periods(exporter_cls, gstin, periods, raw, supplier_names=None):
    """Build over one or more periods; return (file_name, workbook). The mocked lookup
    returns `raw` for each period, so multiple periods exercise the merge."""
    with _mock_names(raw) as stack:
        if supplier_names is not None:
            stack.enter_context(patch.object(GSTR2AExporter, "_supplier_names", lambda self: supplier_names))
        name, content = exporter_cls(gstin, periods).build()
    return name, _load(content)


def _build(exporter_cls, gstin, period, raw, supplier_names=None):
    """Single-period build; return just the workbook."""
    return _build_periods(exporter_cls, gstin, [period], raw, supplier_names)[1]


class TestGSTR2BExport(IntegrationTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.raw = parse_json(read_file(get_data_file_path("test_gstr_2b_v4_0.json")))["data"]
        cls.wb = _build(GSTR2BExporter, GSTIN_2B, PERIOD_2B, cls.raw)

    def test_b2b_row_values(self):
        """Invoice-level row: codes expanded, dates DD/MM/YYYY, period Mon'YY, 0 stays 0."""
        ws = self.wb["B2B"]
        self.assertEqual(ws.cell(7, 1).value, GSTIN_2B)
        self.assertEqual(ws.cell(7, 2).value, "GSTN")
        self.assertEqual(ws.cell(7, 3).value, "S008400")
        self.assertEqual(ws.cell(7, 4).value, "Regular")
        self.assertEqual(ws.cell(7, 5).value, "24/11/2016")
        self.assertEqual(ws.cell(7, 6).value, 729248.16)
        self.assertEqual(ws.cell(7, 7).value, "Haryana")
        self.assertEqual(ws.cell(7, 8).value, "No")
        self.assertEqual(ws.cell(7, 9).value, 12200)
        self.assertEqual(ws.cell(7, 10).value, 183)
        self.assertEqual(ws.cell(7, 11).value, 0)
        self.assertEqual(ws.cell(7, 14).value, "Nov'19")
        self.assertEqual(ws.cell(7, 18).value, "100%")

    def test_itc_available_summary(self):
        """ITC summary from the portal's itcsumm: a section-total row then its detail rows."""
        ws = self.wb["ITC Available"]
        self.assertEqual([ws.cell(9, c).value for c in (4, 5, 6, 7)], [1600, 800, 800, 400])
        self.assertEqual([ws.cell(10, c).value for c in (4, 5, 6, 7)], [400, 200, 200, 100])

    def test_readme_header(self):
        ws = self.wb["Read me"]
        self.assertEqual(ws.cell(4, 3).value, "2019-20")
        self.assertEqual(ws.cell(5, 3).value, "March")
        self.assertEqual(ws.cell(6, 3).value, GSTIN_2B)
        self.assertEqual(ws.cell(7, 3).value, MOCK_LEGAL)
        self.assertEqual(ws.cell(8, 3).value, MOCK_TRADE)
        self.assertEqual(ws.cell(9, 3).value, "14/04/2020")

    def test_all_template_sheets_retained(self):
        self.assertIn("ISDA", self.wb.sheetnames)
        self.assertIn("ECOA(Rejected)", self.wb.sheetnames)

    def test_matches_golden_workbook(self):
        """Every sheet's cell values match the committed golden (values, not bytes)."""
        golden = openpyxl.load_workbook(get_data_file_path(GOLDEN_2B))
        self.assertEqual(self.wb.sheetnames, golden.sheetnames)
        for sheet in golden.sheetnames:
            self.assertEqual(
                _cells(self.wb[sheet]), _cells(golden[sheet]), msg=f"cell mismatch in sheet {sheet!r}"
            )

    def test_multi_period_merges_and_names_file(self):
        """Two periods: document sections concatenate, itcsumm sums, the Read me reflects
        the last period, and the file name joins the periods."""
        single_b2b = _data_row_count(self.wb["B2B"])
        single_revb2b = _data_row_count(self.wb["B2B (ITC Reversal)"])
        single_nonrevsup = self.wb["ITC Available"].cell(9, 4).value

        name, wb = _build_periods(GSTR2BExporter, GSTIN_2B, [PERIOD_2B, "042020"], self.raw)

        self.assertEqual(_data_row_count(wb["B2B"]), 2 * single_b2b)
        # itcrev arrives list-wrapped; both periods' reversal docs must survive the merge
        self.assertEqual(_data_row_count(wb["B2B (ITC Reversal)"]), 2 * single_revb2b)
        self.assertEqual(wb["ITC Available"].cell(9, 4).value, 2 * single_nonrevsup)
        self.assertEqual(wb["Read me"].cell(5, 3).value, "April")
        self.assertEqual(wb["Read me"].cell(4, 3).value, "2020-21")
        self.assertIn(f"{PERIOD_2B}_042020", name)

    def test_import_amendment_split(self):
        """Bills of entry split by isamd: original -> IMPG, amended ('Y') -> IMPGA with the
        amendment type. The payload has no separate impga section."""
        raw = {
            "impg": [
                {"boenum": "BE-ORIG", "isamd": "N", "txval": 100, "igst": 18, "cess": 0},
                {"boenum": "BE-AMD", "isamd": "Y", "amendType": "A", "txval": 200, "igst": 36, "cess": 0},
            ]
        }
        wb = _build(GSTR2BExporter, GSTIN_2B, PERIOD_2B, raw)
        impg, impga = wb["IMPG"], wb["IMPGA"]
        self.assertEqual(impg.cell(7, 3).value, "BE-ORIG")
        self.assertIsNone(impg.cell(8, 3).value)  # amended row is not in IMPG
        self.assertEqual(impga.cell(7, 3).value, "BE-AMD")
        self.assertEqual(impga.cell(7, 6).value, 36)  # amount of tax | integrated
        self.assertEqual(impga.cell(7, 8).value, "Amendment")  # type of amendment (A)

    def test_impgsez_amendment_split(self):
        """SEZ bills of entry are nested under a supplier; split each supplier's boe list."""
        raw = {
            "impgsez": [
                {
                    "ctin": "29ABCDE1234F1Z5",
                    "trdnm": "SEZ Co",
                    "boe": [
                        {"boenum": "S-ORIG", "isamd": "N", "txval": 10, "igst": 2},
                        {"boenum": "S-AMD", "isamd": "Y", "amendType": "G", "txval": 20, "igst": 4},
                    ],
                }
            ]
        }
        wb = _build(GSTR2BExporter, GSTIN_2B, PERIOD_2B, raw)
        self.assertEqual(wb["IMPGSEZ"].cell(7, 1).value, "29ABCDE1234F1Z5")
        self.assertEqual(wb["IMPGSEZ"].cell(7, 5).value, "S-ORIG")
        self.assertEqual(wb["IMPGSEZA"].cell(7, 5).value, "S-AMD")
        self.assertEqual(wb["IMPGSEZA"].cell(7, 10).value, "Addition")  # type of amendment (G)

    def test_itc_reversal_sums_items(self):
        """ITC-reversal invoices carry tax only in items; the sheet shows the summed
        invoice-level taxable value and tax (one row per document)."""
        raw = {
            "itcrev": {
                "b2b": [
                    {
                        "ctin": GSTIN_2B,
                        "trdnm": "X",
                        "inv": [
                            {
                                "inum": "REV-1",
                                "items": [
                                    {"txval": 100, "igst": 18, "cgst": 0, "sgst": 0, "cess": 0},
                                    {"txval": 50, "igst": 9, "cgst": 0, "sgst": 0, "cess": 0},
                                ],
                            }
                        ],
                    }
                ]
            }
        }
        ws = _build(GSTR2BExporter, GSTIN_2B, PERIOD_2B, raw)["B2B (ITC Reversal)"]
        self.assertEqual(ws.cell(7, 3).value, "REV-1")
        self.assertEqual(ws.cell(7, 9).value, 150)  # taxable value (summed)
        self.assertEqual(ws.cell(7, 10).value, 27)  # integrated tax (18 + 9)

    def test_itc_reduction_and_remarks_columns(self):
        """IMS ITC-reduction fields and remarks now map (blank when the payload omits them)."""
        raw = {
            "b2ba": [
                {
                    "ctin": GSTIN_2B,
                    "trdnm": "X",
                    "inv": [
                        {
                            "inum": "A-1",
                            "oinum": "O-1",
                            "itcRedReq": "Y",
                            "declIgst": 5,
                            "declCgst": 2,
                            "declSgst": 2,
                            "declCess": 1,
                            "remarks": "note",
                        }
                    ],
                }
            ]
        }
        ws = _build(GSTR2BExporter, GSTIN_2B, PERIOD_2B, raw)["B2BA"]
        self.assertEqual(ws.cell(8, 16).value, "Yes")  # whether itc to be reduced
        self.assertEqual([ws.cell(8, c).value for c in (17, 18, 19, 20)], [5, 2, 2, 1])
        self.assertEqual(ws.cell(8, 21).value, "note")

    def test_dnra_shifted_tax_columns(self):
        """B2B-DNRA's portal headers mis-merge Taxable Value / Tax Amount a column left;
        the summed taxable value and four taxes must still land in the right cells."""
        raw = {
            "itcrev": {
                "cdnra": [
                    {
                        "ctin": GSTIN_2B,
                        "trdnm": "X",
                        "nt": [
                            {
                                "ntnum": "DN-1",
                                "items": [
                                    {"txval": 100, "igst": 18, "cgst": 0, "sgst": 0, "cess": 0},
                                    {"txval": 100, "igst": 18, "cgst": 0, "sgst": 0, "cess": 0},
                                ],
                            }
                        ],
                    }
                ]
            }
        }
        ws = _build(GSTR2BExporter, GSTIN_2B, PERIOD_2B, raw)["B2B-DNRA"]
        self.assertEqual(ws.cell(8, 6).value, "DN-1")  # revised note number
        self.assertEqual(ws.cell(8, 14).value, 200)  # taxable value
        self.assertEqual(ws.cell(8, 15).value, 36)  # integrated tax
        self.assertEqual(ws.cell(8, 18).value, 0)  # cess


class TestGSTR2AExport(IntegrationTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.wb = _build(GSTR2AExporter, GSTIN_2A, PERIOD_2A, RAW_2A, supplier_names={})

    def test_item_rows_total_and_separator(self):
        """One row per item, a per-invoice total row (number+"-Total", rate "-", taxes
        summed), then a blank separator. Codes/dates verbatim; missing taxes coerced to 0."""
        ws = self.wb["B2B"]
        self.assertEqual(ws.cell(7, 3).value, "INV-001")
        self.assertEqual(ws.cell(7, 5).value, "10-05-2024")
        self.assertEqual([ws.cell(7, c).value for c in (9, 10, 11, 12, 13)], [18, 10000, 0, 900, 900])
        self.assertEqual([ws.cell(8, c).value for c in (9, 10, 11, 12, 13)], [5, 2000, 100, 0, 0])
        self.assertEqual(ws.cell(9, 3).value, "INV-001-Total")
        self.assertEqual(ws.cell(9, 9).value, "-")
        self.assertEqual([ws.cell(9, c).value for c in (10, 11, 12, 13)], [12000, 100, 900, 900])
        self.assertIsNone(ws.cell(10, 1).value)
        self.assertIsNone(ws.cell(10, 3).value)

    def test_isd_document_routing(self):
        """A single docnum/docdt routes to the invoice or credit-note column by doc type;
        the other pair stays blank."""
        ws = self.wb["ISD"]
        self.assertEqual(ws.cell(7, 4).value, "ISD")
        self.assertEqual(ws.cell(7, 5).value, "S0080")
        self.assertEqual(ws.cell(7, 6).value, "03-03-2016")
        self.assertIsNone(ws.cell(7, 7).value)
        self.assertIsNone(ws.cell(7, 8).value)
        self.assertEqual(ws.cell(8, 4).value, "ISDCN")
        self.assertIsNone(ws.cell(8, 5).value)
        self.assertIsNone(ws.cell(8, 6).value)
        self.assertEqual(ws.cell(8, 7).value, "CN-1")
        self.assertEqual(ws.cell(8, 8).value, "04-03-2016")

    def test_b2ba_amendment_columns(self):
        """Amendment sheets label the amended period / amendment type differently from the
        base sheets; `aspd` and `atyp` must still land in those columns (verbatim, like 2A)."""
        raw = {
            "b2ba": [
                {
                    "ctin": GSTIN_2A,
                    "trdnm": "Acme",
                    "inv": [
                        {
                            "inum": "INV-9",
                            "oinum": "INV-1",
                            "aspd": "Apr-24",
                            "atyp": "R",
                            "itms": [{"itm_det": {"rt": 18, "txval": 100, "iamt": 18}}],
                        }
                    ],
                }
            ]
        }
        ws = _build(GSTR2AExporter, GSTIN_2A, PERIOD_2A, raw, supplier_names={})["B2BA"]
        row = next(r for r in range(7, ws.max_row + 1) if ws.cell(r, 6).value == "INV-9")
        self.assertEqual(ws.cell(row, 22).value, "R")  # amendment made, if any
        self.assertEqual(ws.cell(row, 23).value, "Apr-24")  # original tax period reported

    def test_tcs_computed_and_group_fields(self):
        """TCS: operator name from the group, tax period is the export period, and value of
        supplies returned is computed as gross - net-liable."""
        ws = self.wb["TCS"]
        self.assertEqual(ws.cell(7, 1).value, "27AAECS1234F1Z5")
        self.assertEqual(ws.cell(7, 2).value, MOCK_LEGAL)  # operator name = registered legal name
        self.assertEqual(ws.cell(7, 3).value, PERIOD_2A)
        self.assertEqual(ws.cell(7, 4).value, 1000)
        self.assertEqual(ws.cell(7, 5).value, 100)
        self.assertEqual(ws.cell(7, 6).value, 900)

    def test_impgsez_supplier_from_sgstin(self):
        """SEZ imports carry the supplier as sgstin/tdname, not ctin/trdnm."""
        raw = {
            "impgsez": [
                {
                    "sgstin": "29ABCDE1234F1Z5",
                    "tdname": "SEZ Supplier",
                    "benum": "BE9",
                    "txval": 500,
                    "iamt": 90,
                }
            ]
        }
        ws = _build(GSTR2AExporter, GSTIN_2A, PERIOD_2A, raw, supplier_names={})["IMPG SEZ"]
        self.assertEqual(ws.cell(7, 1).value, "29ABCDE1234F1Z5")
        self.assertEqual(ws.cell(7, 2).value, "SEZ Supplier")
        self.assertEqual(ws.cell(7, 5).value, "BE9")  # bill of entry number

    def test_matches_golden_workbook(self):
        """Regression net for 2A: every sheet's cell values match the committed golden
        (built from the inline test payload)."""
        golden = openpyxl.load_workbook(get_data_file_path(GOLDEN_2A))
        self.assertEqual(self.wb.sheetnames, golden.sheetnames)
        for sheet in golden.sheetnames:
            self.assertEqual(
                _cells(self.wb[sheet]), _cells(golden[sheet]), msg=f"cell mismatch in sheet {sheet!r}"
            )

    def test_multi_period_merges_invoice_blocks(self):
        """Two periods: 2A invoice blocks (item rows + total row) repeat across the merged
        sections, and the file name joins the periods."""
        single_rows = _data_row_count(self.wb["B2B"])

        name, wb = _build_periods(GSTR2AExporter, GSTIN_2A, [PERIOD_2A, "062024"], RAW_2A, supplier_names={})

        self.assertEqual(_data_row_count(wb["B2B"]), 2 * single_rows)
        self.assertIn(f"{PERIOD_2A}_062024", name)


class TestPeriodGrouping(IntegrationTestCase):
    """`group_by` decides how many periods share a workbook. Groups follow the financial
    year (April-March), so they never straddle one even when the range does."""

    FY_2024_25: ClassVar[list] = ["042024", "052024", "062024", "072024", "082024", "092024"]

    def test_monthly_gives_one_group_per_period(self):
        self.assertEqual(_group_periods(self.FY_2024_25, "monthly"), [[p] for p in self.FY_2024_25])

    def test_quarterly_splits_on_fiscal_quarters(self):
        self.assertEqual(
            _group_periods(self.FY_2024_25, "quarterly"),
            [["042024", "052024", "062024"], ["072024", "082024", "092024"]],
        )

    def test_half_yearly_and_yearly_keep_the_range_together(self):
        self.assertEqual(_group_periods(self.FY_2024_25, "half_yearly"), [self.FY_2024_25])
        self.assertEqual(_group_periods(self.FY_2024_25, "yearly"), [self.FY_2024_25])

    def test_groups_never_straddle_a_financial_year(self):
        """Feb-May is one calendar quarter-and-a-bit but two fiscal quarters: Feb-Mar closes
        2023-24, Apr-May opens 2024-25. Merging them would report across two years."""
        self.assertEqual(
            _group_periods(["022024", "032024", "042024", "052024"], "quarterly"),
            [["022024", "032024"], ["042024", "052024"]],
        )

    def test_yearly_splits_across_financial_years(self):
        self.assertEqual(
            _group_periods(["032024", "042024"], "yearly"),
            [["032024"], ["042024"]],
        )

    def test_groups_come_back_in_chronological_order(self):
        self.assertEqual(
            _group_periods(["062024", "012025", "042024"], "monthly"),
            [["042024"], ["062024"], ["012025"]],
        )

    def test_all_clubs_every_period_into_one_group(self):
        self.assertEqual(
            _group_periods(["042024", "032024"], "all"),
            [["032024", "042024"]],
        )

    def test_unknown_grouping_falls_back_to_monthly(self):
        self.assertEqual(_group_periods(["042024", "052024"], "weekly"), [["042024"], ["052024"]])


class TestGroupedExport(IntegrationTestCase):
    """Several groups arrive as a zip of full portal-format workbooks, one per group."""

    STEM_2A = f"GSTR-2a-{GSTIN_2A}"  # ReturnType.GSTR2A.value is "GSTR2a", hence the lowercase

    def _export(self, periods, group_by, raw=RAW_2A):
        with _mock_names(raw), patch.object(GSTR2AExporter, "_supplier_names", lambda self: {}):
            return build_export(GSTIN_2A, "GSTR-2A", periods, group_by)

    def test_single_group_stays_a_bare_workbook(self):
        file_name, content = self._export(["042024", "052024", "062024"], "quarterly")

        self.assertTrue(file_name.endswith(".xlsx"))
        self.assertIn("042024_062024", file_name)
        self.assertIn("B2B", _load(content).sheetnames)

    def test_multiple_groups_are_zipped_one_workbook_each(self):
        file_name, content = self._export(["042024", "052024", "062024"], "monthly")

        self.assertTrue(file_name.endswith(".zip"))
        with ZipFile(BytesIO(content)) as archive:
            names = archive.namelist()
            self.assertEqual(
                names,
                [f"{self.STEM_2A}-{period}.xlsx" for period in ("042024", "052024", "062024")],
            )
            # each entry is a real workbook, not a stub
            ws = _load(archive.read(names[0]))["B2B"]
            self.assertEqual(ws.cell(7, 1).value, GSTIN_2A)

    def test_empty_groups_are_skipped_not_fatal(self):
        """One unsynced month must not sink the rest of the range."""
        raw_by_period = {"042024": RAW_2A, "052024": None}

        with (
            patch(
                f"{EXPORT_MODULE}.get_raw_return_data",
                side_effect=lambda gstin, return_type, period: raw_by_period[period],
            ),
            patch.object(GovReturnExporter, "_gstin_info", staticmethod(lambda gstin: {})),
            patch.object(GSTR2AExporter, "_supplier_names", lambda self: {}),
        ):
            file_name, content = build_export(GSTIN_2A, "GSTR-2A", ["042024", "052024"], "monthly")

        self.assertTrue(file_name.endswith(".zip"))
        with ZipFile(BytesIO(content)) as archive:
            self.assertEqual(archive.namelist(), [f"{self.STEM_2A}-042024.xlsx"])

    def test_all_groups_empty_still_throws(self):
        with (
            patch(f"{EXPORT_MODULE}.get_raw_return_data", return_value=None),
            self.assertRaises(frappe.ValidationError),
        ):
            build_export(GSTIN_2A, "GSTR-2A", ["042024", "052024"], "monthly")


class TestSheetRowSpill(IntegrationTestCase):
    """A section with more rows than Excel allows continues on "<sheet> Part N"."""

    def test_overflow_continues_on_a_part_sheet(self):
        # data starts at row 7, so this leaves room for 2 rows per sheet
        with patch(f"{EXPORT_MODULE}.EXCEL_MAX_ROW", 8):
            wb = _build(GSTR2AExporter, GSTIN_2A, PERIOD_2A, RAW_2A, supplier_names={})

        self.assertIn("B2B Part 2", wb.sheetnames)

        first, second = wb["B2B"], wb["B2B Part 2"]
        # 2A B2B here is 4 rows (2 item rows, a total row, a blank separator)
        self.assertEqual(first.cell(7, 1).value, GSTIN_2A)
        self.assertEqual(first.cell(8, 1).value, GSTIN_2A)
        self.assertEqual(second.cell(7, 1).value, GSTIN_2A)

        # the continuation carries the template's headers and merges, not just values
        self.assertEqual(_cells(first)[(5, 1)], _cells(second)[(5, 1)])
        self.assertEqual(len(first.merged_cells.ranges), len(second.merged_cells.ranges))

    def test_no_part_sheet_when_everything_fits(self):
        wb = _build(GSTR2AExporter, GSTIN_2A, PERIOD_2A, RAW_2A, supplier_names={})
        self.assertNotIn("B2B Part 2", wb.sheetnames)


class TestGSTR2ExportIntegration(IntegrationTestCase):
    """End-to-end: store the payload in GST Return Log, then build via the public entry point."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.raw = parse_json(read_file(get_data_file_path("test_gstr_2b_v4_0.json")))["data"]
        store_raw_return_data(GSTIN_2B, ReturnType.GSTR2B.value, PERIOD_2B, cls.raw)

    @patch.object(
        GovReturnExporter,
        "_gstin_info",
        staticmethod(
            lambda gstin: {
                "business_name": MOCK_BUSINESS,
                "legal_name": MOCK_LEGAL,
                "trade_name": MOCK_TRADE,
            }
        ),
    )
    def test_build_export_from_stored_payload(self):
        file_name, content = build_export(GSTIN_2B, "GSTR-2B", [PERIOD_2B])
        self.assertTrue(file_name.endswith(".xlsx"))
        self.assertIn(GSTIN_2B, file_name)

        ws = _load(content)["B2B"]
        self.assertEqual(ws.cell(7, 1).value, GSTIN_2B)
        self.assertEqual(ws.cell(7, 3).value, "S008400")


class TestSupplierNamesAreBatched(IntegrationTestCase):
    """2A resolves a name per supplier. That must be one query for the whole payload, not one
    round trip each — a large return has tens of thousands of suppliers."""

    SUPPLIERS: ClassVar[list] = [f"24AACT{i:04d}F1Z5" for i in range(5)]

    def setUp(self):
        for gstin in self.SUPPLIERS:
            frappe.get_doc({"doctype": "GSTIN", "gstin": gstin, "legal_name": f"LEGAL {gstin}"}).insert(
                ignore_permissions=True
            )

    def tearDown(self):
        for gstin in self.SUPPLIERS:
            frappe.db.delete("GSTIN", {"gstin": gstin})

    def _raw(self):
        return {
            "b2b": [
                {"ctin": gstin, "inv": [{"inum": f"INV-{gstin}", "idt": "10-05-2024", "val": 100}]}
                for gstin in self.SUPPLIERS
            ],
            # SEZ imports key the supplier as sgstin, so the batch must look there too
            "impgsez": [{"sgstin": self.SUPPLIERS[0], "benum": "1", "bedt": "10-05-2024"}],
        }

    def test_registry_is_read_once_not_per_supplier(self):
        per_gstin = Mock(return_value={})
        with (
            patch(f"{EXPORT_MODULE}.get_raw_return_data", return_value=self._raw()),
            patch.object(GSTR2AExporter, "_supplier_names", lambda self: {}),
            patch.object(GovReturnExporter, "_gstin_info", staticmethod(per_gstin)),
        ):
            _, content = GSTR2AExporter(GSTIN_2A, [PERIOD_2A]).build()

        ws = _load(content)["B2B"]
        self.assertEqual(ws.cell(7, 2).value, f"LEGAL {self.SUPPLIERS[0]}")

        # the company's own GSTIN is still looked up once for the Read me header; no supplier
        # should be, or the batch missed them and we're back to a round trip each
        looked_up = {call.args[0] for call in per_gstin.call_args_list}
        self.assertEqual(looked_up - {GSTIN_2A}, set())

    def test_handles_a_large_supplier_set_in_one_query(self):
        """The IN list is deliberately unchunked, so a payload far larger than any real return
        must still go through as a single statement without tripping max_allowed_packet."""
        many = [f"24AACT{i:04d}F1Z5" for i in range(12_000)]
        docdata = {"b2b": [{"ctin": g} for g in many]}

        statements = []
        real = frappe.db.sql

        def counting(query, *a, **k):
            statements.append(query)
            return real(query, *a, **k)

        with patch.object(frappe.db, "sql", counting):
            names = GSTR2AExporter._registry_names(GSTR2AExporter.__new__(GSTR2AExporter), docdata)

        self.assertEqual(len(statements), 1, "supplier lookup was split across statements")
        # only the seeded rows exist, but all 12k were queried without error
        self.assertEqual(len(names), len(self.SUPPLIERS))


class TestExportFileLifecycle(IntegrationTestCase):
    """Generated exports are scratch files: handed over on download, swept if abandoned."""

    def _make_export_file(self, creation=None):
        file = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": f"GSTR-2B-{GSTIN_2B}-{PERIOD_2B}.xlsx",
                "attached_to_doctype": DOCTYPE,
                "attached_to_name": DOCTYPE,
                "is_private": 1,
                "content": b"not-a-real-xlsx",
            }
        ).insert(ignore_permissions=True)

        if creation:
            frappe.db.set_value("File", file.name, "creation", creation, update_modified=False)

        return file.name

    def test_download_streams_the_file(self):
        file_id = self._make_export_file()

        with patch("frappe.utils.response.send_private_file") as send:
            download_export_file(file_id)

        self.assertTrue(send.called, "file was never handed to the browser")
        self.assertTrue(send.call_args.args[0].startswith("/files/"), send.call_args)

        frappe.delete_doc("File", file_id, ignore_permissions=True, delete_permanently=True)

    def test_download_rejects_a_file_this_tool_did_not_create(self):
        """`file_id` is user-supplied, so this guard is all that stands between the endpoint
        and arbitrary private Files."""
        other = frappe.get_doc(
            {"doctype": "File", "file_name": "someone-elses.txt", "is_private": 1, "content": b"x"}
        ).insert(ignore_permissions=True)

        with self.assertRaises(frappe.PermissionError):
            download_export_file(other.name)

        frappe.delete_doc("File", other.name, ignore_permissions=True, delete_permanently=True)

    def test_download_rejects_a_public_file(self):
        """A public File's url has no /private prefix — refuse it rather than IndexError."""
        public = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": "public-export.xlsx",
                "attached_to_doctype": DOCTYPE,
                "attached_to_name": DOCTYPE,
                "is_private": 0,
                "content": b"x",
            }
        ).insert(ignore_permissions=True)

        with self.assertRaises(frappe.PermissionError):
            download_export_file(public.name)

        frappe.delete_doc("File", public.name, ignore_permissions=True, delete_permanently=True)

    def test_sweep_drops_only_abandoned_files(self):
        stale = self._make_export_file(creation=add_days(now_datetime(), -3))
        fresh = self._make_export_file()

        delete_stale_export_files()

        self.assertFalse(frappe.db.exists("File", stale))
        self.assertTrue(frappe.db.exists("File", fresh), "swept a file the user may still download")

        frappe.delete_doc("File", fresh, ignore_permissions=True, delete_permanently=True)


class TestGSTINNameLookup(IntegrationTestCase):
    """`_gstin_info` runs once per unique supplier on a 2A export, so the public API must be
    the last resort: registry hit first, then the production-API gate, then the call."""

    LOOKUP = "india_compliance.gst_india.utils.gstin_info._get_gstin_info"

    def tearDown(self):
        frappe.db.delete("GSTIN", {"gstin": GSTIN_2A})

    def _registry_row(self, **names):
        frappe.get_doc({"doctype": "GSTIN", "gstin": GSTIN_2A, **names}).insert(ignore_permissions=True)

    def test_registry_hit_skips_the_api(self):
        self._registry_row(legal_name=MOCK_LEGAL, trade_name=MOCK_TRADE)

        with (
            patch(f"{EXPORT_MODULE}.is_production_api_enabled", return_value=True),
            patch(self.LOOKUP) as lookup,
        ):
            info = GovReturnExporter._gstin_info(GSTIN_2A)

        self.assertEqual(info.legal_name, MOCK_LEGAL)
        self.assertEqual(info.trade_name, MOCK_TRADE)
        lookup.assert_not_called()

    def test_nameless_registry_row_still_falls_through(self):
        """Rows created before names were cached carry status only — they must not shadow
        the API, or those GSTINs would export blank names forever."""
        self._registry_row(status="Active")

        with (
            patch(f"{EXPORT_MODULE}.is_production_api_enabled", return_value=True),
            patch(self.LOOKUP, return_value={"legal_name": MOCK_LEGAL}) as lookup,
        ):
            self.assertEqual(GovReturnExporter._gstin_info(GSTIN_2A), {"legal_name": MOCK_LEGAL})

        lookup.assert_called_once()

    def test_skipped_when_production_api_is_off(self):
        with (
            patch(f"{EXPORT_MODULE}.is_production_api_enabled", return_value=False),
            patch(self.LOOKUP) as lookup,
        ):
            self.assertEqual(GovReturnExporter._gstin_info(GSTIN_2A), {})

        lookup.assert_not_called()
