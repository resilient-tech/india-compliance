# Copyright (c) 2019, Frappe Technologies Pvt. Ltd. and Contributors
# See license.txt

import json
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook
import frappe
from frappe.tests import IntegrationTestCase, UnitTestCase, change_settings
from frappe.utils import getdate

from india_compliance.gst_india.doctype.gstr_3b_report.gstr_3b_report import (
    GSTR3BExcelExporter,
    download_gstr3b_as_excel,
    format_values,
    make_json,
    view_report,
)
from india_compliance.gst_india.utils.tests import (
    create_purchase_invoice,
    create_sales_invoice,
)

MODULE = "india_compliance.gst_india.doctype.gstr_3b_report.gstr_3b_report"


class TestGSTR3BReport(IntegrationTestCase):
    def setUp(self):
        frappe.set_user("Administrator")
        filters = {"company": "_Test Indian Registered Company"}

        self.maxDiff = None
        for doctype in ("Sales Invoice", "Purchase Invoice", "GSTR 3B Report"):
            frappe.db.delete(doctype, filters=filters)

    @classmethod
    def tearDownClass(cls):
        frappe.db.rollback()

    @change_settings("GST Settings", {"enable_overseas_transactions": 1})
    def test_gstr_3b_report(self):
        month_number_mapping = {
            1: "January",
            2: "February",
            3: "March",
            4: "April",
            5: "May",
            6: "June",
            7: "July",
            8: "August",
            9: "September",
            10: "October",
            11: "November",
            12: "December",
        }

        gst_settings = frappe.get_cached_doc("GST Settings")
        gst_settings.round_off_gst_values = 0
        gst_settings.save()

        create_sales_invoices()
        create_purchase_invoices()

        today = getdate()
        ret_period = f"{today.month:02}{today.year}"

        report = frappe.get_doc(
            {
                "doctype": "GSTR 3B Report",
                "company": "_Test Indian Registered Company",
                "company_gstin": "24AAQCA8719H1ZC",
                "year": today.year,
                "month_or_quarter": month_number_mapping.get(today.month),
            }
        ).insert()

        output = json.loads(report.json_output)

        self.assertDictEqual(
            output,
            {
                "gstin": "24AAQCA8719H1ZC",
                "ret_period": ret_period,
                # 3.1
                "sup_details": {
                    "isup_rev": {
                        "camt": 9.0,
                        "csamt": 0.0,
                        "iamt": 0.0,
                        "samt": 9.0,
                        "txval": 100.0,
                    },
                    "osup_det": {
                        "camt": 18.0,
                        "csamt": 0.0,
                        "iamt": 37.98,
                        "samt": 18.0,
                        "txval": 532.0,
                    },
                    "osup_nil_exmp": {"txval": 100.0},
                    "osup_nongst": {"txval": 222.0},
                    "osup_zero": {"csamt": 0.0, "iamt": 99.9, "txval": 999.0},
                },
                # 3.1.1
                "eco_dtls": {
                    "eco_sup": {
                        "txval": 0,
                        "iamt": 0,
                        "camt": 0,
                        "samt": 0,
                        "csamt": 0,
                    },
                    "eco_reg_sup": {"txval": 100},
                },
                # 3.2
                "inter_sup": {
                    "comp_details": [{"iamt": 18.0, "pos": "29", "txval": 100.0}],
                    "uin_details": [],
                    "unreg_details": [{"iamt": 19.98, "pos": "06", "txval": 111.0}],
                },
                # 4
                "itc_elg": {
                    "itc_avl": [
                        {
                            "camt": 0.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 0.0,
                            "ty": "IMPG",
                        },
                        {
                            "camt": 0.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 0.0,
                            "ty": "IMPS",
                        },
                        {
                            "camt": 9.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 9.0,
                            "ty": "ISRC",
                        },
                        {
                            "camt": 0.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 0.0,
                            "ty": "ISD",
                        },
                        {
                            "camt": 31.5,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 31.5,
                            "ty": "OTH",
                        },
                    ],
                    "itc_inelg": [
                        {
                            "camt": 0.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 0.0,
                            "ty": "RUL",
                        },
                        {
                            "camt": 0.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 0.0,
                            "ty": "OTH",
                        },
                    ],
                    "itc_net": {"camt": 40.5, "csamt": 0.0, "iamt": 0.0, "samt": 40.5},
                    "itc_rev": [
                        {
                            "camt": 0.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 0.0,
                            "ty": "RUL",
                        },
                        {
                            "camt": 0.0,
                            "csamt": 0.0,
                            "iamt": 0.0,
                            "samt": 0.0,
                            "ty": "OTH",
                        },
                    ],
                },
                # 5
                "inward_sup": {
                    "isup_details": [
                        {"inter": 100.0, "intra": 0.0, "ty": "GST"},
                        {"inter": 0.0, "intra": 0.0, "ty": "NONGST"},
                    ]
                },
            },
        )

        exporter = GSTR3BExcelExporter(output)
        response = frappe._dict()

        with patch.object(frappe.local, "response", response, create=True):
            exporter.generate_excel()

        self.assertEqual(response.type, "binary")

    def test_gst_rounding(self):
        gst_settings = frappe.get_cached_doc("GST Settings")
        gst_settings.round_off_gst_values = 1
        gst_settings.save()

        si = create_sales_invoice(
            rate=216,
            is_in_state=True,
            do_not_submit=True,
        )

        # Check for 39 instead of 38.88
        self.assertEqual(si.taxes[0].base_tax_amount_after_discount_amount, 19)

        gst_settings.round_off_gst_values = 1
        gst_settings.save()


class TestGSTR3BReportUnit(UnitTestCase):
    def test_format_values_rounds_nested_numbers(self):
        values = {
            "amount": 12.3456,
            "sections": [{"iamt": 1.235, "nested": [2.678, {"csamt": 9.876}]}],
        }

        self.assertEqual(
            format_values(values),
            {
                "amount": 12.35,
                "sections": [{"iamt": 1.24, "nested": [2.68, {"csamt": 9.88}]}],
            },
        )

    @patch(f"{MODULE}.frappe.get_value", return_value='{"gstin": "24AAQCA8719H1ZC"}')
    @patch(f"{MODULE}.frappe.has_permission")
    def test_view_report_returns_json(self, permission_mock, get_value_mock):
        self.assertEqual(
            view_report("GSTR3B-TEST"),
            {"gstin": "24AAQCA8719H1ZC"},
        )

        permission_mock.assert_called_once_with("GSTR 3B Report", throw=True)
        get_value_mock.assert_called_once_with(
            "GSTR 3B Report", "GSTR3B-TEST", "json_output"
        )

    @patch(f"{MODULE}.frappe.get_value", return_value='{"gstin": "24AAQCA8719H1ZC"}')
    @patch(f"{MODULE}.frappe.has_permission")
    def test_make_json_sets_download_response(self, permission_mock, get_value_mock):
        response = frappe._dict()

        with patch.object(frappe.local, "response", response, create=True):
            make_json("GSTR3B-TEST")

        self.assertEqual(response.filename, "GST3B.json")
        self.assertEqual(response.filecontent, '{"gstin": "24AAQCA8719H1ZC"}')
        self.assertEqual(response.type, "download")
        permission_mock.assert_called_once_with("GSTR 3B Report", throw=True)
        get_value_mock.assert_called_once_with(
            "GSTR 3B Report", "GSTR3B-TEST", "json_output"
        )

    @patch(f"{MODULE}.GSTR3BExcelExporter")
    @patch(
        f"{MODULE}.frappe.get_value",
        return_value='{"gstin": "24AAQCA8719H1ZC", "ret_period": "042024"}',
    )
    @patch(f"{MODULE}.frappe.has_permission")
    def test_download_gstr3b_as_excel_exports_report(
        self, permission_mock, get_value_mock, exporter_mock
    ):
        download_gstr3b_as_excel("GSTR3B-TEST")

        permission_mock.assert_called_once_with("GSTR 3B Report", throw=True)
        get_value_mock.assert_called_once_with(
            "GSTR 3B Report", "GSTR3B-TEST", "json_output"
        )
        exporter_mock.assert_called_once_with(
            {"gstin": "24AAQCA8719H1ZC", "ret_period": "042024"}
        )
        exporter_mock.return_value.generate_excel.assert_called_once_with()

    @patch(f"{MODULE}.frappe.get_value", return_value="")
    @patch(f"{MODULE}.frappe.has_permission")
    def test_download_gstr3b_as_excel_requires_generated_data(
        self, permission_mock, get_value_mock
    ):
        with self.assertRaisesRegex(
            frappe.ValidationError, "Report data not found. Please generate the report."
        ):
            download_gstr3b_as_excel("GSTR3B-TEST")

        permission_mock.assert_called_once_with("GSTR 3B Report", throw=True)
        get_value_mock.assert_called_once_with(
            "GSTR 3B Report", "GSTR3B-TEST", "json_output"
        )

    @patch(f"{MODULE}.os.path.exists")
    def test_generate_excel_throws_when_template_is_missing(self, exists_mock):
        exists_mock.return_value = False

        with self.assertRaisesRegex(
            frappe.ValidationError, "GSTR 3B Excel template not found"
        ):
            GSTR3BExcelExporter({"gstin": "24AAQCA8719H1ZC"}).generate_excel()

    def test_update_worksheet_maps_report_sections_to_template_cells(self):
        workbook = Workbook()
        workbook.active.title = GSTR3BExcelExporter.WORKSHEET_NAME

        exporter = GSTR3BExcelExporter(
            {
                "gstin": "24AAQCA8719H1ZC",
                "ret_period": "042024",
                "sup_details": {
                    "osup_det": {"txval": 100, "iamt": 18, "camt": 9, "csamt": 1},
                    "osup_zero": {"txval": 50, "iamt": 5, "csamt": 0.5},
                    "osup_nil_exmp": {"txval": 10},
                    "isup_rev": {"txval": 20, "iamt": 2, "camt": 1, "csamt": 0.2},
                    "osup_nongst": {"txval": 5},
                },
                "eco_dtls": {"eco_reg_sup": {"txval": 7}},
                "inter_sup": {
                    "unreg_details": [
                        {"pos": "06", "txval": 10, "iamt": 1.8},
                        {"pos": "06", "txval": 5, "iamt": 0.9},
                    ],
                    "comp_details": [{"pos": "29", "txval": 20, "iamt": 3.6}],
                    "uin_details": [{"pos": "00", "txval": 2, "iamt": 0.36}],
                },
                "itc_elg": {
                    "itc_avl": [
                        {"ty": "IMPG", "iamt": 4, "csamt": 0.5},
                        {"ty": "IMPS", "iamt": 3},
                        {"ty": "ISRC", "iamt": 2, "camt": 1, "csamt": 0.2},
                        {"ty": "ISD", "iamt": 1, "camt": 0.5, "csamt": 0.1},
                        {"ty": "OTH", "iamt": 6, "camt": 3, "csamt": 0.3},
                        {"ty": "IGNORED", "iamt": 99},
                    ],
                    "itc_rev": [
                        {"ty": "RUL", "iamt": 1, "camt": 0.5, "csamt": 0.05},
                        {"ty": "OTH", "iamt": 0.5, "camt": 0.25, "csamt": 0.02},
                    ],
                },
                "inward_sup": {
                    "isup_details": [
                        {"ty": "GST", "inter": 8, "intra": 4},
                        {"ty": "NONGST", "inter": 1, "intra": 2},
                        {"ty": "IGNORED", "inter": 9, "intra": 9},
                    ]
                },
            }
        )

        exporter._update_worksheet(SimpleNamespace(wb=workbook))

        worksheet = workbook[GSTR3BExcelExporter.WORKSHEET_NAME]
        self.assertEqual(worksheet.cell(5, 3).value, "24AAQCA8719H1ZC")
        self.assertEqual(worksheet.cell(5, 7).value, "2024-25")
        self.assertEqual(worksheet.cell(6, 7).value, "April")
        self.assertEqual(worksheet.cell(11, 3).value, 100.0)
        self.assertEqual(worksheet.cell(11, 4).value, 18.0)
        self.assertEqual(worksheet.cell(11, 5).value, 9.0)
        self.assertEqual(worksheet.cell(11, 7).value, 1.0)
        self.assertEqual(worksheet.cell(12, 3).value, 50.0)
        self.assertEqual(worksheet.cell(12, 4).value, 5.0)
        self.assertEqual(worksheet.cell(12, 7).value, 0.5)
        self.assertEqual(worksheet.cell(23, 3).value, 7.0)
        self.assertEqual(worksheet.cell(88, 2).value, "00-Other Territory")
        self.assertEqual(worksheet.cell(88, 7).value, 2.0)
        self.assertEqual(worksheet.cell(88, 8).value, 0.36)
        self.assertEqual(worksheet.cell(89, 2).value, "06-Haryana")
        self.assertEqual(worksheet.cell(89, 3).value, 15.0)
        self.assertEqual(worksheet.cell(89, 4).value, 2.7)
        self.assertEqual(worksheet.cell(90, 2).value, "29-Karnataka")
        self.assertEqual(worksheet.cell(90, 5).value, 20.0)
        self.assertEqual(worksheet.cell(90, 6).value, 3.6)
        self.assertEqual(worksheet.cell(31, 3).value, 4.0)
        self.assertEqual(worksheet.cell(31, 6).value, 0.5)
        self.assertEqual(worksheet.cell(33, 3).value, 2.0)
        self.assertEqual(worksheet.cell(33, 4).value, 1.0)
        self.assertEqual(worksheet.cell(33, 6).value, 0.2)
        self.assertEqual(worksheet.cell(37, 3).value, 1.0)
        self.assertEqual(worksheet.cell(37, 4).value, 0.5)
        self.assertEqual(worksheet.cell(37, 6).value, 0.05)
        self.assertEqual(worksheet.cell(48, 4).value, 8.0)
        self.assertEqual(worksheet.cell(48, 5).value, 4.0)
        self.assertEqual(worksheet.cell(49, 4).value, 1.0)
        self.assertEqual(worksheet.cell(49, 5).value, 2.0)

    def test_set_cell_ignores_merged_cells(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = GSTR3BExcelExporter.WORKSHEET_NAME
        worksheet.merge_cells("B2:C2")

        exporter = GSTR3BExcelExporter({})
        exporter.worksheet = worksheet
        exporter._set_cell(2, 3, "ignored")

        self.assertIsNone(worksheet["B2"].value)


def create_sales_invoices():
    create_sales_invoice(is_in_state=True)
    create_sales_invoice(
        customer="_Test Registered Composition Customer",
        is_out_state=True,
    )
    create_sales_invoice(
        customer="_Test Unregistered Customer",
        is_in_state=True,
    )
    # Unregistered Out of state
    create_sales_invoice(
        customer="_Test Unregistered Customer",
        is_out_state=True,
        place_of_supply="06-Haryana",
        rate=111,
    )
    # Same Item Nil-Rated
    create_sales_invoice(item_tax_template="Nil-Rated - _TIRC")
    # Non Gst item
    create_sales_invoice(item_code="_Test Non GST Item", rate=222)
    # Zero Rated
    create_sales_invoice(
        customer_address="_Test Registered Customer-Billing-1",
        is_export_with_gst=False,
        rate=444,
    )
    create_sales_invoice(
        customer_address="_Test Registered Customer-Billing-1",
        is_export_with_gst=True,
        is_out_state=True,
        rate=555,
    )
    # E-commerce reverse charge
    create_sales_invoice(
        customer="_Test Registered Customer",
        is_reverse_charge=True,
        item_code="_Test Trading Goods 1",
        rate=100,
        ecommerce_gstin="29AABCF8078M1C8",
        is_in_state_rcm=True,
    )
    # Reverse Charge Sales
    create_sales_invoice(
        customer="_Test Registered Customer",
        is_reverse_charge=True,
        item_code="_Test Trading Goods 1",
        rate=121,
        is_in_state_rcm=True,
    )


def create_purchase_invoices():
    create_purchase_invoice(is_in_state=True)
    create_purchase_invoice(rate=250, qty=1, is_in_state=True)
    create_purchase_invoice(supplier="_Test Registered Composition Supplier")
    create_purchase_invoice(
        is_in_state_rcm=True,
        supplier="_Test Unregistered Supplier",
        is_reverse_charge=True,
    )
